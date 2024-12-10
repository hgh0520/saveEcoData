#API key : test_efe36b31c30950ae6ce720e3fda13a6cd620aaa23398292d37bc7e837be265fdefe8d04e6d233bd35cf2fabdeb93fb0d
npcshtop_addr = "https://open.api.nexon.com/mabinogi/v1/npcshop/list?" # 상점 카탈로그 조회
auction_list_addr =  "https://open.api.nexon.com/mabinogi/v1/auction/list?" # 경매장 매물 검색
auction_history_addr = "https://open.api.nexon.com/mabinogi/v1/auction/history?" # 경매장 거래 내역 조회
horn_addr = "https://open.api.nexon.com/mabinogi/v1/horn-bugle-world/history?" # 거뿔 내역 조회

import requests
import json

from datetime import datetime
import pytz

import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import platform

file_path = " "

# OS 구분
if platform.system() == "Darwin":
    print("This is macOS.")
    file_path = "/Users/hankwanghyun/Google Drive/내 드라이브/Mabinogi/에코스톤_판매내역_search.xlsx"
elif platform.system() == "Windows":
    print("This is Windows.")
    file_path = "D:\Google 드라이브\Mabinogi\에코스톤_판매내역_search.xlsx"  # file path for windows
else:
    print(f"This is an unsupported OS: {platform.system()}")
    exit()

ecoNames = []
ecoPrice = []
ecoTimes = []
ecoLevel = []
ecoStats = []
ecoAwake = []
ecoBound = []

headers = {
    "x-nxopen-api-key": "test_efe36b31c30950ae6ce720e3fda13a6cd620aaa23398292d37bc7e837be265fdefe8d04e6d233bd35cf2fabdeb93fb0d"
}

def get_response(urlString):
    response = requests.get(urlString, headers = headers)
    return response

def check_json_key(json, key):
    try:
        buf = json[key]
    except KeyError:
        return False
    return True

def utc_to_kst(utc_time):

    # print(datetime.strptime(utc_time, "%Y-%m-%dT%H:%M:%S.000Z"))
    
    utc_time  = datetime.strptime(utc_time, "%Y-%m-%dT%H:%M:%S.000Z")
    utc_time  = pytz.utc.localize(utc_time )

    kst = pytz.timezone('Asia/Seoul')
    kst_time = utc_time.astimezone(kst).strftime('%Y-%m-%d %H:%M:%S')

    # print(kst_time.strftime('%Y-%m-%d %H:%M:%S'))
    # print(f"한국 시간: {kst_time.strftime('%Y-%m-%d %H:%M:%S')}")

    return kst_time

def auto_adjust_column_width(ws):
    # 각 열의 최대 길이를 구하고 그에 맞게 열 너비를 설정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 열 이름 (A, B, C 등)
        for cell in col:
            try:
                # 셀 값이 문자열일 경우 길이 계산
                cell_value = str(cell.value)
                # 한글 문자는 길이를 두 배로 계산
                if any('\uac00' <= char <= '\ud7af' for char in cell_value):  # 한글이 포함되어 있으면
                    max_length = max(max_length, len(cell_value) * 2)
                else:
                    max_length = max(max_length, len(cell_value))
            except:
                pass
        adjusted_width = (max_length + 2)  # 조금 여유를 두고 설정
        ws.column_dimensions[column].width = adjusted_width

def format_number(value):
    if isinstance(value, (int, float)):
        if value >= 100000000:  # 억 단위
            man = (value % 100000000) // 10000  # 억 단위 나머지
            eok = value // 100000000  # 억 단위
            if man > 0:
                return f"{eok}억 {man}만"
            return f"{eok}억"
        elif value >= 10000:  # 만 단위
            return f"{value // 10000}만"
        elif value >= 1000:  # 천 단위
            return f"{value // 1000}K"  # 천 단위는 K로
        else:
            return value  # 천 미만은 그대로 출력
    return value

def get_auction_history():
    auction_item_category="에코스톤"
    urlString = auction_history_addr + "auction_item_category=" + auction_item_category

    # print(urlString)

    res_data = get_response(urlString).json()
    # response = requests.get(urlString, headers = headers)

    # print(res_data)

    # res_data = response.json()

    # print(res_data)

    auction_history = res_data["auction_history"]

    # print(auction_history)

    # print(len(auction_history))


    
    posLevel = 0
    posStats = 1
    posAwake = 2
    posBound = 3

    for i in range(len(auction_history)):
        if len(auction_history[i]["item_option"]) == 4:
            # print(auction_history[i]["item_option"][posLevel]["option_value"])  # get ecostone level 1~30
            # print(auction_history[i]["item_option"][posStats]["option_value"])  # get ecostone stats 1~130
            # print(auction_history[i]["item_option"][posAwake]["option_value"])  # get ecostone awake type and level
            # print(auction_history[i]["item_option"][posBound]["option_value"])  # get ecostone remain presonalized bound

            ecoNames.append(auction_history[i]["item_name"])
            ecoPrice.append(auction_history[i]["auction_price_per_unit"])
            ecoTimes.append(utc_to_kst(auction_history[i]["date_auction_buy"]))

            ecoLevel.append(auction_history[i]["item_option"][posLevel]["option_value"])
            ecoStats.append(auction_history[i]["item_option"][posStats]["option_value"])
            ecoAwake.append(auction_history[i]["item_option"][posAwake]["option_value"])
            ecoBound.append(auction_history[i]["item_option"][posBound]["option_value"])

            
                    

    for i in range(len(ecoNames)):
        ecoAwake[i] = ecoAwake[i] + "(" + ecoBound[i] + ")"
        # print(ecoNames[i], ecoPrice[i], ecoTimes[i], ecoLevel[i], ecoStats[i], ecoAwake[i]) 
    


# 현재 시간 가져오기
now = datetime.now()

# 형식화된 현재 시간 출력
last_update_time = now.strftime("%Y-%m-%d %H:%M:%S")
print("마지막 업데이트 시간:", last_update_time)

get_auction_history()

data = {
    "에코스톤 타입": ecoNames,
    "에코스톤 등급": ecoLevel,
    "에코스톤 고유 능력": ecoStats,
    "에코스톤 각성 능력(전해)": ecoAwake,
    "판매 시각": ecoTimes,
    "판매 가격": ecoPrice,
    "마지막 업데이트 시각": [None] * len(ecoPrice)
}

df = pd.DataFrame(data)

# print(df)

data_sheet = "TotalData"
search_sheet = "Browse"  # 새 시트 이름

if os.path.exists(file_path) and not df.empty:
    # 기존 데이터 읽기
    existing_df = pd.read_excel(file_path, sheet_name = data_sheet)

    # 'Time' 컬럼을 datetime 형식으로 변환
    existing_df['판매 시각'] = pd.to_datetime(existing_df['판매 시각'])
    df['판매 시각'] = pd.to_datetime(df['판매 시각'])

    # 새로운 데이터의 마지막 시간과 기존 데이터의 첫 번째 시간 비교
    last_new_data_time = df['판매 시각'].iloc[-1]  # 새로운 데이터의 마지막 시간
    first_existing_data_time = existing_df['판매 시각'].iloc[0]  # 기존 데이터의 첫 번째 시간
    
    # print(last_new_data_time, " vs", first_existing_data_time)

    while last_new_data_time <= first_existing_data_time:
        # 마지막 데이터가 과거이면, 새로운 데이터의 마지막 -1을 비교
        
        df = df.iloc[:-1]  # 마지막 행을 제거
        if not df.empty:
            last_new_data_time = df['판매 시각'].iloc[-1]  # 새로운 데이터의 마지막 시간 갱신
        else:
            break  # 더 이상 비교할 데이터가 없으면 종료
    
    print(df)

    if not df.empty:
        existing_df['마지막 업데이트 시각'] = None

        # 기존 데이터를 2번째 행부터 시작하도록 재정렬
        updated_df = pd.concat([existing_df.iloc[:0], df, existing_df.iloc[0:]], ignore_index=True)

        # 엑셀 파일에 저장
        with pd.ExcelWriter(file_path, engine="openpyxl", mode = 'a', if_sheet_exists='replace') as writer:
            updated_df.to_excel(writer, index=0, sheet_name=data_sheet)

            # 엑셀 워크북과 워크시트 접근
            wb = writer.book
            ws = wb[data_sheet]

            for row in ws.iter_rows(min_col=6, max_col=6, min_row=2, max_row=ws.max_row):  # F열은 6번째 열
                for cell in row:
                    cell.value = format_number(cell.value)  # 셀 값에 단위 적용
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws['G2'] = last_update_time
            ws['G3'] = " "

            # 열 너비 자동 조정
            auto_adjust_column_width(ws)
    else:
        with pd.ExcelWriter(file_path, engine="openpyxl", mode = 'a', if_sheet_exists='replace') as writer:
            # 엑셀 워크북과 워크시트 접근
            wb = writer.book
            ws = wb[data_sheet]

            ws['G2'] = last_update_time
            ws['G3'] = " "
    
elif not os.path.exists(file_path):
    # 파일이 없으면 새로 생성
    df.to_excel(file_path, index=False)

    # 엑셀 파일을 생성한 후 열 너비 자동 조정
    wb = load_workbook(file_path)
    ws_data = wb.active
    ws_data.title = data_sheet
    ws_data = wb[data_sheet]
    
    for row in ws_data.iter_rows(min_col=6, max_col=6, min_row=2, max_row=ws_data.max_row):  # F열은 6번째 열
        for cell in row:
            cell.value = format_number(cell.value)  # 셀 값에 단위 적용

    for row in ws_data.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 열 너비 자동 조정
    auto_adjust_column_width(ws_data)

    # 검색 시트 이름
    ws_search = wb.create_sheet(title = search_sheet, index = 0)
    ws_search['A2'] = '각성 능력 검색'
    ws_search['D2'] = '검색하려는 각성 능력을 입력해주세요'
    
    ws_search['A5'] = '에코스톤 타입'
    ws_search['B5'] = '에코스톤 등급'
    ws_search['C5'] = '에코스톤 고유 능력'
    ws_search['D5'] = '에코스톤 각성 능력(전해)'
    ws_search['E5'] = '판매 시각'
    ws_search['F5'] = '판매 가격'

    # 열 너비 자동 조정
    auto_adjust_column_width(ws_search)
    
    ws_search['A5'] = '=IF(ISBLANK(B3), " ", IFERROR(FILTER(TotalData!A:F, ISNUMBER(SEARCH(Browse!B3, TotalData!D:D))), "검색 결과가 없습니다"))'
    #=FILTER(TotalData!A:F, ISNUMBER(SEARCH(Search!B1, TotalData!D:D)), "일치하는 데이터가 없습니다.")

    # 파일 저장
    wb.save(file_path)

else:
    print("There is nothing to update.")